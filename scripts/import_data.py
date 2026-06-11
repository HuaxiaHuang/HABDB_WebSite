from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

from app.config import settings
from app.database import Base, engine, SessionLocal
from app import models


def clean(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def slugify(v, prefix="item"):
    s = re.sub(r"[^A-Za-z0-9]+", "-", (v or "").strip()).strip("-").lower()
    return s or prefix


def norm_species(v):
    return re.sub(r"\s+", " ", (v or "").replace("_", " ")).strip().lower()


def display_species(v):
    return re.sub(r"\s+", " ", (v or "").replace("_", " ")).strip()


def first(row, names):
    lower = {k.lower(): v for k, v in row.items()}
    for n in names:
        if n in row and row[n]:
            return row[n]
        if n.lower() in lower and lower[n.lower()]:
            return lower[n.lower()]
    return ""


def read_sheet(path: Path, sheet=None, limit=None):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    header = None
    out = []
    for raw in ws.iter_rows(values_only=True):
        vals = [clean(v) for v in raw]
        if not any(vals):
            continue
        if header is None:
            header = vals
            continue
        out.append({header[i] if i < len(header) and header[i] else f"col_{i+1}": vals[i] if i < len(vals) else "" for i in range(max(len(header), len(vals)))})
        if limit and len(out) >= limit:
            break
    return out


def read_all_sheets(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    return {ws.title: read_sheet(path, ws.title) for ws in wb.worksheets}


def count_fasta(path: Path):
    if not path.exists():
        return 0
    count = 0
    with path.open("rb") as f:
        for line in f:
            if line.startswith(b">"):
                count += 1
    return count


def iter_fasta(path: Path):
    sid = desc = ""
    chunks = []
    if not path.exists():
        return
    with path.open("r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith(">"):
                if sid:
                    yield sid, desc, "".join(chunks)
                desc = line[1:]
                sid = desc.split()[0]
                chunks = []
            else:
                chunks.append(line)
        if sid:
            yield sid, desc, "".join(chunks)


def sha256(path: Path, max_size=1024 * 1024 * 1024):
    if not path.exists():
        return "missing"
    if path.stat().st_size > max_size:
        return "skipped-large-file"
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def toxic_label(v):
    val = str(v or "").strip().lower()
    if val in {"1", "yes", "true", "toxic"}:
        return "toxic"
    if val in {"0", "no", "false", "non-toxic", "nontoxic"}:
        return "harmful non-toxic"
    return "unknown"


def import_all(source: Path, reset: bool):
    if reset:
        Base.metadata.drop_all(bind=engine)
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()
    try:
        if not reset and db.query(models.Species).first():
            print("Existing data detected; use --reset to rebuild.")
            return
        print(f"Importing HABDB resources from {source}")
        species_rows = read_sheet(source / "HABDB-List.xlsx")
        marker_sheets = read_all_sheets(source / "HABDB-AS" / "HABDB-AS-18SrRNA" / "HABDB-AS-18SrRNA-information.xlsx")
        rep_rows = marker_sheets.get("Representative_Seqs", [])
        comprehensive_rows = marker_sheets.get("Comprehensive_Seqs", [])
        plus_rows = read_sheet(source / "HABDB-AS-18SrRNA-plus.xlsx")
        genome_rows = read_sheet(source / "HABDB-AS-Genomeinfo.xlsx")

        rep_by = {norm_species(first(r, ["Species"])): r for r in rep_rows if first(r, ["Species"])}
        extended = Counter(norm_species(first(r, ["Species"])) for r in comprehensive_rows + plus_rows if first(r, ["Species"]))
        cds_by = {}
        cds_dir = source / "HABDB-AS" / "HABDB-AS-CDS" / "HAB-CDS"
        for f in cds_dir.glob("*_with_NCBI_sequences_Tagged.xlsx") if cds_dir.exists() else []:
            try:
                wb = load_workbook(f, read_only=True, data_only=True)
                n = max(wb.worksheets[0].max_row - 1, 0)
            except Exception:
                n = 0
            cds_by[norm_species(f.name.replace("_with_NCBI_sequences_Tagged.xlsx", ""))] = n

        canonical = {}
        row_by = {}
        for r in species_rows:
            name = first(r, ["Species"])
            if name:
                key = norm_species(name)
                canonical[key] = display_species(name)
                row_by[key] = r
        for key in cds_by:
            canonical.setdefault(key, display_species(key))

        species_id_by_key = {}
        for i, key in enumerate(sorted(canonical, key=lambda k: canonical[k].lower()), 1):
            r = row_by.get(key, {})
            rep = rep_by.get(key, {})
            sid = f"HABDB-AS-{i:04d}-{slugify(canonical[key])}"
            species_id_by_key[key] = sid
            obj = models.Species(
                id=sid,
                name=canonical[key],
                taxid=first(r, ["NCBI taxid"]) or first(rep, ["NCBI taxid"]),
                domain=first(r, ["Domain", "Domin"]),
                kingdom=first(r, ["Kingdom"]),
                phylum=first(r, ["Phylum"]),
                class_name=first(r, ["Class"]),
                order=first(r, ["Order"]),
                family=first(r, ["Family"]),
                genus=first(r, ["Genus"]),
                toxic_status=toxic_label(first(r, ["Toxic"]) or first(rep, ["Toxic"])),
                toxin_class=first(r, ["Toxin class", "Toxin"]) or "unknown",
                representative_18srrna=first(r, ["Accession number"]) or first(rep, ["18S rRNA"]),
                sequence_status=first(r, ["Sequence Status"]) or first(rep, ["Sequence Status"]),
                extended_18srrna_count=extended.get(key, 0),
                genome_count=0,
                cds_count=cds_by.get(key, 0),
                source="HABDB-List.xlsx; HABDB-AS-18SrRNA-information.xlsx; HABDB-AS-Genomeinfo.xlsx",
            )
            db.merge(obj)
        db.commit()
        print(f"Imported species: {len(canonical)}")

        for r in comprehensive_rows[:5000]:
            name = first(r, ["Species"])
            acc = first(r, ["Sequence_ID", "Accession number"])
            if not name or not acc:
                continue
            key = norm_species(name)
            seq = first(r, ["Sequence"])
            db.merge(models.MarkerSequence(
                species_id=species_id_by_key.get(key),
                species_name=display_species(name),
                marker_type="18SrRNA",
                accession=acc,
                description=first(r, ["Description"]),
                length_bp=int(first(r, ["Length(bp)", "Length"]) or len(seq) or 0),
                sequence=seq,
            ))
        db.commit()

        genome_counts = Counter()
        for r in genome_rows:
            name = first(r, ["Genome"])
            key = norm_species(name)
            db.add(models.Genome(
                species_id=species_id_by_key.get(key),
                genome=display_species(name),
                strain=first(r, ["Strain"]),
                assembly_accession=first(r, ["Assembly accession"]),
                assembly_level=first(r, ["Assembly level"]),
                gc_percent=first(r, ["GC percent"]),
                scaffold_n50=first(r, ["Scaffold N50"]),
                contig_n50=first(r, ["Contig N50"]),
            ))
            if species_id_by_key.get(key):
                genome_counts[species_id_by_key[key]] += 1
        for sid, count in genome_counts.items():
            sp = db.get(models.Species, sid)
            if sp:
                sp.genome_count = count
        db.commit()

        import_genes(db, source)
        import_downloads(db, source)
        write_stats(db, source)
        write_manifest(db, source)
        db.commit()
        print("Import completed.")
    finally:
        db.close()


def infer_gene(row, module):
    text = " ".join(str(v) for v in row.values())
    for p in ["sxtA", "sxtB", "sxtG", "sxtH", "sxtI", "sxtN", "sxtO", "dabA", "dabB", "dabC", "dabD", "lcf", "lbp", "luciferase"]:
        if re.search(p, text, re.I):
            return p
    return first(row, ["Gene", "Gene family", "Name", "Protein", "Annotation"]) or module


def import_genes(db, source):
    files = [("STX/PST", "HABDB-FG-STX-SeedSequence.xlsx"), ("Domoic acid", "HABDB-FG-DA-SeedSequence.xlsx"), ("Bioluminescence", "HABDB-FG-Bioluminescence-SeedSequence.xlsx")]
    fam = defaultdict(lambda: {"seed": 0, "files": set(), "annotation": ""})
    for module, fn in files:
        for i, row in enumerate(read_sheet(source / fn), 1):
            gene = infer_gene(row, module)
            fid = f"{slugify(module)}-{slugify(gene)}"
            seq = first(row, ["Sequence", "Protein sequence", "AA sequence"])
            fam[(module, gene)]["seed"] += 1
            fam[(module, gene)]["files"].add(fn)
            fam[(module, gene)]["annotation"] = fam[(module, gene)]["annotation"] or first(row, ["Annotation", "Description", "Protein", "Function"]) or gene
            db.merge(models.FunctionalSequence(
                id=f"HABDB-FG-{slugify(module)}-{slugify(gene)}-{i:05d}",
                gene_family_id=fid,
                module=module,
                gene_family=gene,
                accession=first(row, ["Accession", "Accession number", "Protein ID", "Entry", "ID"]),
                annotation=fam[(module, gene)]["annotation"],
                source_database=first(row, ["Source", "Database", "DB"]) or "curated seed",
                evidence="seed",
                database_layer="seed",
                sequence_length=len(re.sub(r"\s+", "", seq)),
                sequence=seq,
            ))
    full_counts = Counter()
    idmap = source / "HABs_FuncDB_id2genemap.txt"
    if idmap.exists():
        for line in idmap.read_text(encoding="utf-8", errors="replace").splitlines():
            parts = re.split(r"\t|,|\s+", line.strip())
            if len(parts) >= 2:
                full_counts[parts[-1]] += 1
    for (module, gene), info in fam.items():
        full = sum(c for g, c in full_counts.items() if g.lower() == gene.lower() or g.lower() in gene.lower())
        db.merge(models.GeneFamily(
            id=f"{slugify(module)}-{slugify(gene)}",
            module=module,
            gene_family=gene,
            annotation=info["annotation"],
            seed_count=info["seed"],
            full_count=full,
            source_files=";".join(sorted(info["files"])),
        ))
    db.commit()
    print(f"Imported gene families: {len(fam)}")


def import_downloads(db, source):
    patterns = ["HABDB-List.xlsx", "HABDB-AS-18SrRNA-plus.xlsx", "HABDB-AS-Genomeinfo.xlsx", "HABDB-FG-*-SeedSequence.xlsx", "HABs_FuncDB_Full.database", "HABs_FuncDB_id2genemap.txt", "HABs_Func_db.dmnd", "HABDB-AS/HABDB-AS-18SrRNA/HABs_18S_sequences.fasta", "HABDB-AS/HABDB-AS-18SrRNA/HABs_Final_Kraken2_db.tar.gz", "HABDB-AS/HABDB-AS-Genome/fna.zip"]
    files = []
    for p in patterns:
        files.extend(source.glob(p))
    for f in files:
        rel = str(f.relative_to(source)).replace("\\", "/")
        category = "Metadata"
        if f.suffix.lower() in {".fasta", ".database", ".dmnd", ".gz"}:
            category = "Sequence search"
        if f.suffix.lower() == ".zip":
            category = "Genome archive"
        db.merge(models.DownloadFile(
            id=slugify(f.name),
            name=f.name,
            relative_path=rel,
            category=category,
            format=f.suffix.lower().lstrip("."),
            size_bytes=f.stat().st_size,
            checksum_sha256=sha256(f),
            description=f"HABDB downloadable {category.lower()} resource. 18SrRNA resources use the unified 18SrRNA naming convention.",
        ))
    db.commit()
    print(f"Imported download files: {len(files)}")


def write_stats(db, source):
    stats = {
        "species_count": db.query(models.Species).count(),
        "marker_18srrna_count": db.query(models.MarkerSequence).count(),
        "genome_assemblies": db.query(models.Genome).count(),
        "gene_family_count": db.query(models.GeneFamily).count(),
        "functional_sequence_seed_count": db.query(models.FunctionalSequence).filter(models.FunctionalSequence.database_layer == "seed").count(),
        "functional_gene_full_records": count_fasta(source / "HABs_FuncDB_Full.database"),
        "download_file_count": db.query(models.DownloadFile).count(),
        "generated_at": datetime.utcnow().isoformat(timespec="seconds"),
    }
    for k, v in stats.items():
        db.merge(models.Statistic(key=k, value=str(v)))
    print(json.dumps(stats, indent=2))


def write_manifest(db, source):
    files = [d for d in db.query(models.DownloadFile).all()]
    manifest = {
        "release": settings.habdb_release,
        "generated_at": datetime.utcnow().isoformat(timespec="seconds"),
        "marker_naming": "18SrRNA",
        "files": [{"name": f.name, "relative_path": f.relative_path, "size_bytes": f.size_bytes, "sha256": f.checksum_sha256} for f in files],
    }
    out = settings.habdb_job_dir.parent / "release_manifest.json"; out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    db.add(models.ReleaseManifest(release=settings.habdb_release, manifest_json=json.dumps(manifest, ensure_ascii=False)))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=settings.habdb_dataresource)
    parser.add_argument("--reset", action="store_true")
    args = parser.parse_args()
    import_all(args.source, args.reset)


if __name__ == "__main__":
    main()
